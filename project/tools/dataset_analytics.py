import argparse
import json
import math
import re
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_INPUT = Path(r"C:\flutter\flutter-Project\project\dataset\#5 finish_attraction_enriched.xlsx")
DEFAULT_OUTPUT_DIR = Path(r"C:\flutter\flutter-Project\project\dataset\analytics")


FEATURE_COLUMNS = [
    ("region", "REGION_NAME_TH", "single", "ภูมิภาคของสถานที่"),
    ("province", "PROVINCE_NAME_TH", "single", "จังหวัดของสถานที่"),
    ("category", "ATT_CATEGORY_LABEL", "single", "หมวดหมู่หลักของสถานที่"),
    ("type", "ATT_TYPE_LABEL", "single", "ประเภทสถานที่"),
    ("activity", "ATT_ACTIVITY", "multi", "กิจกรรมหรือความสนใจที่เกี่ยวข้อง"),
]


def clean_text(value: object) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    return " ".join(str(value).strip().split())


def split_activity(value: object) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    return [part.strip() for part in re.split(r"[,/|]", text) if part.strip()]


def split_urls(value: object) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    return [part.strip() for part in text.split("|") if part.strip()]


def autosize(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="5A189A")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    autosize(ws)


def add_table_rows(ws, headers: list[str], rows: list[list[object]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws)


def count_single(df: pd.DataFrame, column: str) -> Counter:
    return Counter(clean_text(value) for value in df[column] if clean_text(value))


def count_activity(df: pd.DataFrame) -> Counter:
    counter: Counter = Counter()
    for value in df["ATT_ACTIVITY"]:
        counter.update(split_activity(value))
    return counter


def make_count_rows(counter: Counter, total_rows: int, limit: int | None = None) -> list[list[object]]:
    items = counter.most_common(limit)
    return [
        [index, name, count, round((count / total_rows) * 100, 2)]
        for index, (name, count) in enumerate(items, start=1)
    ]


def add_bar_chart(ws, title: str, data_rows: int, anchor: str = "F2") -> None:
    if data_rows < 1:
        return
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = "จำนวน"
    chart.x_axis.title = "รายการ"
    data = Reference(ws, min_col=3, min_row=1, max_row=data_rows + 1)
    categories = Reference(ws, min_col=2, min_row=2, max_row=data_rows + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 8
    chart.width = 16
    ws.add_chart(chart, anchor)


def add_pie_chart(ws, title: str, data_rows: int, anchor: str = "F2") -> None:
    if data_rows < 1:
        return
    chart = PieChart()
    chart.title = title
    data = Reference(ws, min_col=3, min_row=2, max_row=data_rows + 1)
    labels = Reference(ws, min_col=2, min_row=2, max_row=data_rows + 1)
    chart.add_data(data)
    chart.set_categories(labels)
    chart.height = 8
    chart.width = 12
    ws.add_chart(chart, anchor)


def write_analytics_workbook(df: pd.DataFrame, output_path: Path) -> dict[str, object]:
    total_rows = len(df)
    region_counter = count_single(df, "REGION_NAME_TH")
    province_counter = count_single(df, "PROVINCE_NAME_TH")
    category_counter = count_single(df, "ATT_CATEGORY_LABEL")
    type_counter = count_single(df, "ATT_TYPE_LABEL")
    activity_counter = count_activity(df)

    image_counts = df["IMAGE_URLS"].map(lambda value: len(split_urls(value))) if "IMAGE_URLS" in df else pd.Series([0] * total_rows)
    youtube_counts = df["YOUTUBE_URLS"].map(lambda value: len(split_urls(value))) if "YOUTUBE_URLS" in df else pd.Series([0] * total_rows)
    rows_with_images = int((image_counts > 0).sum())
    rows_with_youtube = int((youtube_counts > 0).sum())

    summary = {
        "total_attractions": total_rows,
        "regions": len(region_counter),
        "provinces": len(province_counter),
        "categories": len(category_counter),
        "types": len(type_counter),
        "activities": len(activity_counter),
        "rows_with_images": rows_with_images,
        "image_coverage_percent": round((rows_with_images / total_rows) * 100, 2),
        "rows_without_images": total_rows - rows_with_images,
        "rows_with_youtube": rows_with_youtube,
        "youtube_coverage_percent": round((rows_with_youtube / total_rows) * 100, 2),
        "rows_without_youtube": total_rows - rows_with_youtube,
        "top_province": province_counter.most_common(1)[0][0] if province_counter else "",
        "top_category": category_counter.most_common(1)[0][0] if category_counter else "",
        "top_type": type_counter.most_common(1)[0][0] if type_counter else "",
        "top_activity": activity_counter.most_common(1)[0][0] if activity_counter else "",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    summary_rows = [
        ["Metric", "Value", "Description"],
        ["Total attractions", summary["total_attractions"], "จำนวนสถานที่ทั้งหมดใน dataset"],
        ["Regions", summary["regions"], "จำนวนภูมิภาค"],
        ["Provinces", summary["provinces"], "จำนวนจังหวัด"],
        ["Categories", summary["categories"], "จำนวนหมวดหมู่"],
        ["Types", summary["types"], "จำนวนประเภทสถานที่"],
        ["Activities", summary["activities"], "จำนวนกิจกรรม/ความสนใจ"],
        ["Rows with images", summary["rows_with_images"], "จำนวนสถานที่ที่มีรูปภาพ"],
        ["Image coverage (%)", summary["image_coverage_percent"], "เปอร์เซ็นต์สถานที่ที่มีรูปภาพ"],
        ["Rows without images", summary["rows_without_images"], "จำนวนสถานที่ที่ยังไม่มีรูปภาพ"],
        ["Rows with YouTube", summary["rows_with_youtube"], "จำนวนสถานที่ที่มีลิงก์ YouTube"],
        ["YouTube coverage (%)", summary["youtube_coverage_percent"], "เปอร์เซ็นต์สถานที่ที่มี YouTube"],
        ["Rows without YouTube", summary["rows_without_youtube"], "จำนวนสถานที่ที่ยังไม่มี YouTube"],
        ["Top province", summary["top_province"], "จังหวัดที่มีสถานที่มากที่สุด"],
        ["Top category", summary["top_category"], "หมวดหมู่ที่พบมากที่สุด"],
        ["Top type", summary["top_type"], "ประเภทที่พบมากที่สุด"],
        ["Top activity", summary["top_activity"], "กิจกรรมที่พบมากที่สุด"],
    ]
    for row in summary_rows:
        ws.append(row)
    style_sheet(ws)

    media_ws = wb.create_sheet("Media Coverage")
    media_rows = [
        ["No", "Media Status", "Count", "Percent"],
        [1, "มีรูปภาพ", rows_with_images, round((rows_with_images / total_rows) * 100, 2)],
        [2, "ไม่มีรูปภาพ", total_rows - rows_with_images, round(((total_rows - rows_with_images) / total_rows) * 100, 2)],
        [3, "มี YouTube", rows_with_youtube, round((rows_with_youtube / total_rows) * 100, 2)],
        [4, "ไม่มี YouTube", total_rows - rows_with_youtube, round(((total_rows - rows_with_youtube) / total_rows) * 100, 2)],
    ]
    for row in media_rows:
        media_ws.append(row)
    style_sheet(media_ws)
    add_pie_chart(media_ws, "Image Coverage", 2, "F2")

    sheets = [
        ("Region Count", region_counter, None, "จำนวนสถานที่ตามภูมิภาค"),
        ("Province Count", province_counter, 31, "จำนวนสถานที่ตามจังหวัด"),
        ("Category Count", category_counter, None, "จำนวนสถานที่ตามหมวดหมู่"),
        ("Type Count", type_counter, 30, "จำนวนสถานที่ตามประเภท"),
        ("Activity Count", activity_counter, 30, "จำนวนกิจกรรมที่พบใน dataset"),
    ]
    for sheet_name, counter, limit, title in sheets:
        count_ws = wb.create_sheet(sheet_name)
        rows = make_count_rows(counter, total_rows, limit)
        add_table_rows(count_ws, ["No", "Name", "Count", "Percent of attractions"], rows)
        add_bar_chart(count_ws, title, len(rows), "F2")

    missing_ws = wb.create_sheet("Missing Media")
    missing_rows = []
    for index, row in df.iterrows():
        image_count = len(split_urls(row.get("IMAGE_URLS", "")))
        youtube_count = len(split_urls(row.get("YOUTUBE_URLS", "")))
        if image_count == 0 or youtube_count == 0:
            missing_rows.append(
                [
                    int(index) + 2,
                    clean_text(row.get("ATT_ID", "")),
                    clean_text(row.get("ATT_NAME_TH", "")),
                    clean_text(row.get("PROVINCE_NAME_TH", "")),
                    image_count,
                    youtube_count,
                ]
            )
    add_table_rows(
        missing_ws,
        ["Excel Row", "ATT_ID", "ATT_NAME_TH", "Province", "Image Count", "YouTube Count"],
        missing_rows,
    )

    wb.save(output_path)
    return summary


def write_transform_workbook(df: pd.DataFrame, output_path: Path) -> dict[str, object]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Transform Overview"
    overview_rows = [
        ["Feature Group", "Source Column", "Transform Method", "Example Feature Name", "Value Meaning"],
        ["region", "REGION_NAME_TH", "One-hot / binary 0-1", "region_ภาคเหนือ", "1 = อยู่ในภาคเหนือ, 0 = ไม่ใช่"],
        ["province", "PROVINCE_NAME_TH", "One-hot / binary 0-1", "province_เชียงใหม่", "1 = อยู่จังหวัดเชียงใหม่, 0 = ไม่ใช่"],
        ["category", "ATT_CATEGORY_LABEL", "One-hot / binary 0-1", "category_ธรรมชาติ", "1 = หมวดหมู่ธรรมชาติ, 0 = ไม่ใช่"],
        ["type", "ATT_TYPE_LABEL", "One-hot / binary 0-1", "type_น้ำตก", "1 = ประเภทน้ำตก, 0 = ไม่ใช่"],
        ["activity", "ATT_ACTIVITY", "Multi-label binary 0-1", "activity_ถ่ายรูป", "1 = มีกิจกรรมถ่ายรูป, 0 = ไม่มี"],
    ]
    for row in overview_rows:
        ws.append(row)
    style_sheet(ws)

    summary: dict[str, object] = {}
    for feature_group, source_column, mode, description in FEATURE_COLUMNS:
        sheet_name = f"{feature_group.title()} Features"
        feature_ws = wb.create_sheet(sheet_name)
        values: list[str]
        if mode == "multi":
            values = sorted(count_activity(df))
        else:
            values = sorted(count_single(df, source_column))

        rows = [
            [
                index,
                f"{feature_group}_{value}",
                source_column,
                value,
                "1 = มี/ตรงกับค่านี้, 0 = ไม่มี/ไม่ตรง",
                description,
            ]
            for index, value in enumerate(values, start=1)
        ]
        add_table_rows(
            feature_ws,
            ["Code", "Feature Name", "Source Column", "Original Value", "Encoding", "Description"],
            rows,
        )
        summary[feature_group] = len(values)

    sample_ws = wb.create_sheet("Encoded Example")
    sample_headers = [
        "ATT_NAME_TH",
        "REGION_NAME_TH",
        "PROVINCE_NAME_TH",
        "ATT_CATEGORY_LABEL",
        "ATT_TYPE_LABEL",
        "ATT_ACTIVITY",
        "Example Binary Features",
    ]
    sample_rows = []
    for _, row in df.head(20).iterrows():
        activities = split_activity(row.get("ATT_ACTIVITY", ""))
        binary_features = [
            f"region_{clean_text(row.get('REGION_NAME_TH', ''))}=1",
            f"province_{clean_text(row.get('PROVINCE_NAME_TH', ''))}=1",
            f"category_{clean_text(row.get('ATT_CATEGORY_LABEL', ''))}=1",
            f"type_{clean_text(row.get('ATT_TYPE_LABEL', ''))}=1",
        ]
        binary_features.extend(f"activity_{activity}=1" for activity in activities)
        sample_rows.append(
            [
                clean_text(row.get("ATT_NAME_TH", "")),
                clean_text(row.get("REGION_NAME_TH", "")),
                clean_text(row.get("PROVINCE_NAME_TH", "")),
                clean_text(row.get("ATT_CATEGORY_LABEL", "")),
                clean_text(row.get("ATT_TYPE_LABEL", "")),
                ", ".join(activities),
                " | ".join(binary_features),
            ]
        )
    add_table_rows(sample_ws, sample_headers, sample_rows)

    wb.save(output_path)
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description="Create dataset analytics and feature transform report files.")
    parser.add_argument("--input", default=str(DEFAULT_INPUT))
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    args = parser.parse_args()

    input_path = Path(args.input)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    df = pd.read_excel(input_path, dtype=str).fillna("")

    analytics_path = output_dir / "dataset_analytics_summary.xlsx"
    transform_path = output_dir / "feature_transform_table.xlsx"
    summary_json_path = output_dir / "dataset_analytics_summary.json"

    analytics_summary = write_analytics_workbook(df, analytics_path)
    transform_summary = write_transform_workbook(df, transform_path)

    summary = {
        "input": str(input_path),
        "analyticsWorkbook": str(analytics_path),
        "transformWorkbook": str(transform_path),
        "analytics": analytics_summary,
        "transformFeatureCounts": transform_summary,
    }
    summary_json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
