import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def split_urls(value) -> list[str]:
    if not value:
        return []
    return list(dict.fromkeys(url.strip() for url in str(value).split("|") if url.strip()))


def main() -> None:
    parser = argparse.ArgumentParser(description="Merge Excel IMAGE_URLS/YOUTUBE_URLS into dataset/attractions.json")
    parser.add_argument("--excel", required=True)
    parser.add_argument("--json", default="dataset/attractions.json")
    parser.add_argument("--start-row", type=int, default=2)
    parser.add_argument("--end-row", type=int, default=0)
    parser.add_argument("--images", action="store_true")
    parser.add_argument("--youtube", action="store_true")
    args = parser.parse_args()

    workbook = load_workbook(args.excel, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    name_col = headers.index("ATT_NAME_TH")
    image_col = headers.index("IMAGE_URLS") if "IMAGE_URLS" in headers else None
    youtube_col = headers.index("YOUTUBE_URLS") if "YOUTUBE_URLS" in headers else None

    media: dict[str, dict[str, list[str]]] = {}
    end_row = args.end_row or sheet.max_row
    for row in sheet.iter_rows(min_row=args.start_row, max_row=end_row, values_only=True):
        name = str(row[name_col]).strip()
        media[name] = {}
        if args.images and image_col is not None:
            media[name]["images"] = split_urls(row[image_col])
        if args.youtube and youtube_col is not None:
            media[name]["youtubeUrls"] = split_urls(row[youtube_col])

    json_path = Path(args.json)
    data = json.loads(json_path.read_text(encoding="utf-8"))

    updated = 0
    counts: dict[str, dict[int, int]] = {"images": {}, "youtubeUrls": {}}
    for item in data:
        name = str(item.get("nameTh", "")).strip()
        if name not in media:
            continue

        for field, urls in media[name].items():
            item[field] = urls
            if field == "youtubeUrls":
                item["youtubeUrl"] = urls[0] if urls else ""
            counts[field][len(urls)] = counts[field].get(len(urls), 0) + 1
        updated += 1

    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"updated_rows": updated, "counts": counts}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
