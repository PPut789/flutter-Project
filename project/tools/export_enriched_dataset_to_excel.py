import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


SOURCE_EXCEL = Path(
    r"C:\Users\poomp\OneDrive - Rajamangala University of Technology Thanyaburi"
    r"\Project\Chapter 1 2 3\Dataset\#5 finish_attraction.xlsx"
)
SOURCE_JSON = Path(r"C:\flutter\flutter-Project\project\dataset\attractions.json")
OUTPUT_EXCEL = Path(r"C:\flutter\flutter-Project\project\dataset\#5 finish_attraction_enriched.xlsx")
OUTPUT_JSON = Path(r"C:\flutter\flutter-Project\project\dataset\attractions.json")


def join_urls(value: list[str] | None) -> str:
    if not value:
        return ""
    return "|".join(dict.fromkeys(str(url).strip() for url in value if str(url).strip()))


def normalize_name(value: object) -> str:
    return " ".join(str(value or "").split())


def main() -> None:
    parser = argparse.ArgumentParser(description="Export app media data back into an enriched Excel dataset copy.")
    parser.add_argument("--source-excel", default=str(SOURCE_EXCEL))
    parser.add_argument("--source-json", default=str(SOURCE_JSON))
    parser.add_argument("--output-excel", default=str(OUTPUT_EXCEL))
    parser.add_argument("--output-json", default=str(OUTPUT_JSON))
    args = parser.parse_args()

    source_excel = Path(args.source_excel)
    source_json = Path(args.source_json)
    output_excel = Path(args.output_excel)
    output_json = Path(args.output_json)

    output_excel.parent.mkdir(parents=True, exist_ok=True)
    output_json.parent.mkdir(parents=True, exist_ok=True)

    app_data = json.loads(source_json.read_text(encoding="utf-8"))
    media_by_name = {
        normalize_name(item.get("nameTh", "")): {
            "IMAGE_URLS": join_urls(item.get("images")),
            "YOUTUBE_URLS": join_urls(item.get("youtubeUrls")),
        }
        for item in app_data
        if normalize_name(item.get("nameTh", ""))
    }

    workbook = load_workbook(source_excel)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]

    if "ATT_NAME_TH" not in headers:
        raise SystemExit("Missing ATT_NAME_TH column in source Excel.")

    name_col = headers.index("ATT_NAME_TH") + 1
    media_columns = ["IMAGE_URLS", "YOUTUBE_URLS"]
    column_indexes: dict[str, int] = {}

    for column_name in media_columns:
        if column_name in headers:
            column_indexes[column_name] = headers.index(column_name) + 1
        else:
            new_col = sheet.max_column + 1
            sheet.cell(row=1, column=new_col, value=column_name)
            column_indexes[column_name] = new_col
            headers.append(column_name)

    updated = 0
    image_rows = 0
    youtube_rows = 0
    unmatched = 0

    for row_index in range(2, sheet.max_row + 1):
        name = normalize_name(sheet.cell(row=row_index, column=name_col).value)
        media = media_by_name.get(name)
        if not media:
            unmatched += 1
            continue

        for column_name, column_index in column_indexes.items():
            sheet.cell(row=row_index, column=column_index, value=media[column_name])

        if media["IMAGE_URLS"]:
            image_rows += 1
        if media["YOUTUBE_URLS"]:
            youtube_rows += 1
        updated += 1

    workbook.save(output_excel)
    output_json.write_text(json.dumps(app_data, ensure_ascii=False, indent=2), encoding="utf-8")

    print(
        json.dumps(
            {
                "output_excel": str(output_excel),
                "output_json": str(output_json),
                "updated_rows": updated,
                "unmatched_rows": unmatched,
                "rows_with_images": image_rows,
                "rows_with_youtube": youtube_rows,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
