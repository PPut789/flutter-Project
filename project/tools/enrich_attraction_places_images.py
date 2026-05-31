import argparse
import json
import os
import time
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


IMAGE_COLUMN = "IMAGE_URLS"
PLACE_ID_COLUMN = "GOOGLE_PLACE_ID"


def load_env(env_path: Path) -> None:
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def request_json(url: str) -> dict:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "travelthai-places-images/1.0",
            "Accept": "application/json",
        },
    )
    with urllib.request.urlopen(request, timeout=20) as response:
        return json.loads(response.read().decode("utf-8"))


def get_or_create_column(headers: list[str], name: str) -> int:
    if name in headers:
        return headers.index(name) + 1

    headers.append(name)
    return len(headers)


def cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def build_queries(row_values: dict[str, str]) -> list[str]:
    name_th = row_values.get("ATT_NAME_TH", "")
    name_en = row_values.get("ATT_NAME_EN", "")
    province = row_values.get("PROVINCE_NAME_TH", "")
    district = row_values.get("DISTRICT_NAME_TH", "")

    queries = [
        " ".join(part for part in [name_th, district, province, "Thailand"] if part),
        " ".join(part for part in [name_th, province, "Thailand"] if part),
        " ".join(part for part in [name_en, province, "Thailand"] if part),
        " ".join(part for part in [name_en, "Thailand"] if part),
    ]
    return list(dict.fromkeys(query for query in queries if query))


def text_search_place(query: str, api_key: str) -> tuple[str, list[str]]:
    params = {
        "query": query,
        "key": api_key,
        "language": "th",
        "region": "th",
    }
    url = "https://maps.googleapis.com/maps/api/place/textsearch/json?" + urllib.parse.urlencode(
        params,
    )
    data = request_json(url)
    status = data.get("status", "")

    if status not in ("OK", "ZERO_RESULTS"):
        raise RuntimeError(f"Places Text Search status={status}: {data.get('error_message', '')}")

    results = data.get("results", [])
    if not results:
        return "", []

    first = results[0]
    place_id = first.get("place_id", "")
    photo_refs = [
        photo.get("photo_reference", "")
        for photo in first.get("photos", [])
        if photo.get("photo_reference")
    ]
    return place_id, photo_refs


def place_details_photos(place_id: str, api_key: str, max_results: int) -> list[str]:
    params = {
        "place_id": place_id,
        "fields": "photos",
        "key": api_key,
        "language": "th",
    }
    url = "https://maps.googleapis.com/maps/api/place/details/json?" + urllib.parse.urlencode(
        params,
    )
    data = request_json(url)
    status = data.get("status", "")

    if status not in ("OK", "ZERO_RESULTS"):
        raise RuntimeError(f"Place Details status={status}: {data.get('error_message', '')}")

    photos = data.get("result", {}).get("photos", [])
    return [
        photo.get("photo_reference", "")
        for photo in photos[:max_results]
        if photo.get("photo_reference")
    ]


def photo_url(photo_reference: str, api_key: str, max_width: int) -> str:
    params = {
        "maxwidth": max_width,
        "photo_reference": photo_reference,
        "key": api_key,
    }
    return "https://maps.googleapis.com/maps/api/place/photo?" + urllib.parse.urlencode(params)


def resolved_photo_url(photo_reference: str, api_key: str, max_width: int) -> str:
    url = photo_url(photo_reference, api_key, max_width)
    request = urllib.request.Request(url, headers={"User-Agent": "travelthai-places-images/1.0"})
    with urllib.request.urlopen(request, timeout=20) as response:
        return response.url


def enrich_places_images(
    input_path: Path,
    output_path: Path,
    env_path: Path,
    limit: int,
    start_row: int,
    overwrite: bool,
    sleep_seconds: float,
    min_images: int,
    max_images: int,
) -> None:
    load_env(env_path)
    api_key = os.environ.get("GOOGLE_API_KEY", "").strip()
    if not api_key:
        raise SystemExit("Missing GOOGLE_API_KEY in tools/media_enrichment.env")

    workbook = load_workbook(input_path)
    sheet = workbook.active

    headers = [cell_text(cell.value) for cell in sheet[1]]
    image_col = get_or_create_column(headers, IMAGE_COLUMN)
    place_id_col = get_or_create_column(headers, PLACE_ID_COLUMN)
    sheet.cell(row=1, column=image_col).value = IMAGE_COLUMN
    sheet.cell(row=1, column=place_id_col).value = PLACE_ID_COLUMN

    header_to_index = {name: index + 1 for index, name in enumerate(headers)}
    required = ["ATT_NAME_TH", "PROVINCE_NAME_TH"]
    missing = [name for name in required if name not in header_to_index]
    if missing:
        raise SystemExit(f"Missing required columns: {', '.join(missing)}")

    processed = 0
    first_data_row = max(2, start_row)

    for row_number in range(first_data_row, sheet.max_row + 1):
        if limit and processed >= limit:
            break

        row_values = {
            name: cell_text(sheet.cell(row=row_number, column=index).value)
            for name, index in header_to_index.items()
        }
        place_name = row_values.get("ATT_NAME_TH", "")
        if not place_name:
            continue

        current_images = cell_text(sheet.cell(row=row_number, column=image_col).value)
        if current_images and not overwrite:
            continue

        found_place_id = ""
        found_photo_refs: list[str] = []

        try:
            for query in build_queries(row_values):
                found_place_id, found_photo_refs = text_search_place(query, api_key)
                if found_place_id:
                    break

            if found_place_id and len(found_photo_refs) < min_images:
                detail_photo_refs = place_details_photos(
                    found_place_id,
                    api_key,
                    max_images,
                )
                found_photo_refs = list(
                    dict.fromkeys([*found_photo_refs, *detail_photo_refs]),
                )

            urls: list[str] = []
            for photo_reference in found_photo_refs:
                if len(urls) >= max_images:
                    break
                url = resolved_photo_url(photo_reference, api_key, 1200)
                if url not in urls:
                    urls.append(url)
            sheet.cell(row=row_number, column=image_col).value = "|".join(urls)
            sheet.cell(row=row_number, column=place_id_col).value = found_place_id
            processed += 1
            print(
                f"[{processed}] row {row_number}: {place_name} "
                f"place={'yes' if found_place_id else 'no'} images={len(urls)}",
                flush=True,
            )
            workbook.save(output_path)
            time.sleep(sleep_seconds)
        except Exception as error:
            print(f"[error] row {row_number}: {place_name}: {error}", flush=True)
            workbook.save(output_path)
            time.sleep(sleep_seconds)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"Saved: {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Fill IMAGE_URLS from Google Places Photos."
    )
    parser.add_argument("--input", required=True, help="Path to source .xlsx")
    parser.add_argument("--output", required=True, help="Path to output .xlsx")
    parser.add_argument(
        "--env",
        default="tools/media_enrichment.env",
        help="Env file with GOOGLE_API_KEY",
    )
    parser.add_argument("--limit", type=int, default=20, help="Rows to enrich")
    parser.add_argument("--start-row", type=int, default=2, help="Excel row number")
    parser.add_argument("--overwrite", action="store_true", help="Replace existing images")
    parser.add_argument("--sleep", type=float, default=0.2, help="Delay between rows")
    parser.add_argument("--min-images", type=int, default=2, help="Preferred minimum images")
    parser.add_argument("--max-images", type=int, default=8, help="Maximum images per place")
    args = parser.parse_args()

    enrich_places_images(
        input_path=Path(args.input),
        output_path=Path(args.output),
        env_path=Path(args.env),
        limit=args.limit,
        start_row=args.start_row,
        overwrite=args.overwrite,
        sleep_seconds=args.sleep,
        min_images=args.min_images,
        max_images=args.max_images,
    )


if __name__ == "__main__":
    main()
