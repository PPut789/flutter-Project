import argparse
import json
import os
import time
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


IMAGE_COLUMN = "IMAGE_URLS"
YOUTUBE_COLUMN = "YOUTUBE_URLS"


def load_env(env_path: Path) -> None:
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def request_json(url: str) -> dict:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "travelthai-media-enrichment/1.0",
            "Accept": "application/json",
        },
    )
    with urllib.request.urlopen(request, timeout=15) as response:
        return json.loads(response.read().decode("utf-8"))


def google_image_search(query: str, api_key: str, cse_id: str, max_results: int) -> list[str]:
    params = {
        "key": api_key,
        "cx": cse_id,
        "q": query,
        "searchType": "image",
        "imgType": "photo",
        "safe": "active",
        "num": max_results,
    }
    url = "https://www.googleapis.com/customsearch/v1?" + urllib.parse.urlencode(params)
    data = request_json(url)

    image_urls: list[str] = []
    for item in data.get("items", []):
        link = item.get("link", "")
        if link.startswith("http"):
            image_urls.append(link)

    return image_urls[:max_results]


def wikimedia_image_search(query: str, max_results: int) -> list[str]:
    params = {
        "action": "query",
        "generator": "search",
        "gsrsearch": query,
        "gsrnamespace": "6",
        "gsrlimit": max_results,
        "prop": "imageinfo",
        "iiprop": "url",
        "format": "json",
        "origin": "*",
    }
    url = "https://commons.wikimedia.org/w/api.php?" + urllib.parse.urlencode(params)
    data = request_json(url)

    pages = data.get("query", {}).get("pages", {})
    image_urls: list[str] = []
    for page in pages.values():
        image_info = page.get("imageinfo", [])
        if image_info and image_info[0].get("url", "").startswith("http"):
            image_urls.append(image_info[0]["url"])

    return image_urls[:max_results]


def openverse_image_search(query: str, max_results: int) -> list[str]:
    params = {
        "q": query,
        "page_size": max_results,
        "mature": "false",
    }
    url = "https://api.openverse.engineering/v1/images?" + urllib.parse.urlencode(params)
    data = request_json(url)

    image_urls: list[str] = []
    for item in data.get("results", []):
        url_value = item.get("url", "")
        if url_value.startswith("http"):
            image_urls.append(url_value)

    return image_urls[:max_results]


def image_search_with_fallback(
    query: str,
    api_key: str,
    cse_id: str,
    max_results: int,
) -> tuple[list[str], str]:
    try:
        return google_image_search(query, api_key, cse_id, max_results), "google"
    except Exception as error:
        print(f"[google image fallback] {error}", flush=True)

    image_urls = wikimedia_image_search(query, max_results)
    if image_urls:
        return image_urls, "wikimedia"

    return openverse_image_search(query, max_results), "openverse"


def youtube_search(query: str, api_key: str, max_results: int) -> list[str]:
    params = {
        "key": api_key,
        "part": "snippet",
        "q": query,
        "type": "video",
        "maxResults": max_results,
        "regionCode": "TH",
        "relevanceLanguage": "th",
    }
    url = "https://www.googleapis.com/youtube/v3/search?" + urllib.parse.urlencode(params)
    data = request_json(url)

    youtube_urls: list[str] = []
    for item in data.get("items", []):
        video_id = item.get("id", {}).get("videoId")
        if video_id:
            youtube_urls.append(f"https://www.youtube.com/watch?v={video_id}")

    return youtube_urls[:max_results]


def canonical_youtube_url(url: str) -> str:
    parsed = urllib.parse.urlparse(url)
    host = parsed.netloc.lower().replace("www.", "")

    video_id = ""
    if host == "youtu.be":
        video_id = parsed.path.strip("/").split("/")[0]
    elif host.endswith("youtube.com"):
        if parsed.path == "/watch":
            query = urllib.parse.parse_qs(parsed.query)
            video_id = query.get("v", [""])[0]
        elif parsed.path.startswith("/shorts/"):
            video_id = parsed.path.split("/")[2] if len(parsed.path.split("/")) > 2 else ""
        elif parsed.path.startswith("/embed/"):
            video_id = parsed.path.split("/")[2] if len(parsed.path.split("/")) > 2 else ""

    if not video_id:
        return ""

    return f"https://www.youtube.com/watch?v={video_id}"


def youtube_custom_search(
    query: str,
    api_key: str,
    cse_id: str,
    max_results: int,
) -> list[str]:
    params = {
        "key": api_key,
        "cx": cse_id,
        "q": f"{query} site:youtube.com/watch OR site:youtu.be",
        "safe": "active",
        "num": min(10, max_results * 2),
    }
    url = "https://www.googleapis.com/customsearch/v1?" + urllib.parse.urlencode(params)
    data = request_json(url)

    youtube_urls: list[str] = []
    seen: set[str] = set()
    for item in data.get("items", []):
        links = [
            item.get("link", ""),
            item.get("formattedUrl", ""),
            item.get("htmlFormattedUrl", ""),
        ]
        for link in links:
            clean_url = canonical_youtube_url(link)
            if clean_url and clean_url not in seen:
                youtube_urls.append(clean_url)
                seen.add(clean_url)
            if len(youtube_urls) >= max_results:
                return youtube_urls

    return youtube_urls[:max_results]


def is_quota_error(error: Exception) -> bool:
    return "HTTP Error 403" in str(error)


def get_or_create_column(headers: list[str], name: str) -> int:
    if name in headers:
        return headers.index(name) + 1

    headers.append(name)
    return len(headers)


def cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def build_query(row_values: dict[str, str], suffix: str) -> str:
    parts = [
        row_values.get("ATT_NAME_TH", ""),
        row_values.get("PROVINCE_NAME_TH", ""),
        suffix,
    ]
    return " ".join(part for part in parts if part).strip()


def build_image_queries(row_values: dict[str, str]) -> list[str]:
    name_th = row_values.get("ATT_NAME_TH", "")
    name_en = row_values.get("ATT_NAME_EN", "")
    province = row_values.get("PROVINCE_NAME_TH", "")

    queries = [
        " ".join(part for part in [name_th, province, "สถานที่ท่องเที่ยว รูปภาพ"] if part),
        " ".join(part for part in [name_en, "Thailand tourist attraction"] if part),
        " ".join(part for part in [name_th, "Thailand"] if part),
        " ".join(part for part in [name_en, "Thailand"] if part),
    ]
    return list(dict.fromkeys(query for query in queries if query))


def build_youtube_queries(row_values: dict[str, str], mode: str) -> list[str]:
    name_th = row_values.get("ATT_NAME_TH", "")
    name_en = row_values.get("ATT_NAME_EN", "")
    province = row_values.get("PROVINCE_NAME_TH", "")

    thai_queries = [
        " ".join(part for part in [name_th, province, "สถานที่ท่องเที่ยว"] if part),
    ]
    english_queries = [
        " ".join(part for part in [name_en, province, "Thailand travel"] if part),
        " ".join(part for part in [name_en, "Thailand"] if part),
    ]
    mixed_queries = [
        " ".join(part for part in [name_th, "Thailand"] if part),
    ]

    if mode == "thai-only":
        queries = thai_queries
    elif mode == "english-only":
        queries = english_queries
    else:
        queries = [*thai_queries, *english_queries, *mixed_queries]

    return list(dict.fromkeys(query for query in queries if query))


def enrich_workbook(
    input_path: Path,
    output_path: Path,
    env_path: Path,
    limit: int,
    start_row: int,
    overwrite: bool,
    sleep_seconds: float,
    media: str,
    stop_after_quota_errors: int,
    youtube_query_mode: str,
    youtube_provider: str,
) -> None:
    load_env(env_path)

    api_key = os.environ.get("GOOGLE_API_KEY", "").strip()
    cse_id = os.environ.get("GOOGLE_CSE_ID", "").strip()

    if not api_key or not cse_id:
        raise SystemExit(
            "Missing GOOGLE_API_KEY or GOOGLE_CSE_ID. "
            "Copy tools/media_enrichment.env.example to tools/media_enrichment.env and fill both values."
        )

    workbook = load_workbook(input_path)
    sheet = workbook.active

    headers = [cell_text(cell.value) for cell in sheet[1]]
    image_col = get_or_create_column(headers, IMAGE_COLUMN)
    youtube_col = get_or_create_column(headers, YOUTUBE_COLUMN)

    sheet.cell(row=1, column=image_col).value = IMAGE_COLUMN
    sheet.cell(row=1, column=youtube_col).value = YOUTUBE_COLUMN

    header_to_index = {name: index + 1 for index, name in enumerate(headers)}
    required = ["ATT_NAME_TH", "PROVINCE_NAME_TH"]
    missing = [name for name in required if name not in header_to_index]
    if missing:
        raise SystemExit(f"Missing required columns: {', '.join(missing)}")

    processed = 0
    consecutive_quota_errors = 0
    first_data_row = max(2, start_row)

    for row_number in range(first_data_row, sheet.max_row + 1):
        if limit and processed >= limit:
            break

        row_values = {
            name: cell_text(sheet.cell(row=row_number, column=index).value)
            for name, index in header_to_index.items()
        }

        place_name = row_values.get("ATT_NAME_TH", "")
        if not place_name:
            continue

        current_images = cell_text(sheet.cell(row=row_number, column=image_col).value)
        current_youtube = cell_text(sheet.cell(row=row_number, column=youtube_col).value)

        should_fetch_images = media in ("all", "images") and (overwrite or not current_images)
        should_fetch_youtube = media in ("all", "youtube") and (overwrite or not current_youtube)

        if not should_fetch_images and not should_fetch_youtube:
            continue

        image_urls: list[str] = []
        youtube_urls: list[str] = []

        try:
            if should_fetch_images:
                try:
                    image_urls = []
                    image_source = ""
                    for image_query in build_image_queries(row_values):
                        image_urls, image_source = image_search_with_fallback(
                            image_query,
                            api_key,
                            cse_id,
                            4,
                        )
                        if image_urls:
                            break
                    sheet.cell(row=row_number, column=image_col).value = "|".join(image_urls)
                    if image_urls:
                        print(
                            f"[image source] row {row_number}: {image_source}",
                            flush=True,
                        )
                except Exception as error:
                    print(
                        f"[image error] row {row_number}: {place_name}: {error}",
                        flush=True,
                    )

            if should_fetch_youtube:
                try:
                    youtube_urls = []
                    had_youtube_quota_error = False
                    for youtube_query in build_youtube_queries(
                        row_values,
                        youtube_query_mode,
                    ):
                        try:
                            if youtube_provider == "custom-search":
                                youtube_urls = youtube_custom_search(
                                    youtube_query,
                                    api_key,
                                    cse_id,
                                    3,
                                )
                            else:
                                youtube_urls = youtube_search(youtube_query, api_key, 3)
                            if youtube_urls:
                                break
                        except Exception as error:
                            if is_quota_error(error):
                                had_youtube_quota_error = True
                                raise
                            print(
                                f"[youtube query error] row {row_number}: "
                                f"{place_name}: {error}",
                                flush=True,
                            )
                    if youtube_urls or overwrite:
                        sheet.cell(row=row_number, column=youtube_col).value = "|".join(
                            youtube_urls,
                        )
                    if had_youtube_quota_error:
                        consecutive_quota_errors += 1
                except Exception as error:
                    if is_quota_error(error):
                        consecutive_quota_errors += 1
                    print(
                        f"[youtube error] row {row_number}: {place_name}: {error}",
                        flush=True,
                    )

            processed += 1
            print(
                f"[{processed}] row {row_number}: {place_name} "
                f"images={len(image_urls) if should_fetch_images else 'skip'} "
                f"youtube={len(youtube_urls) if should_fetch_youtube else 'skip'}",
                flush=True,
            )
            workbook.save(output_path)
            if consecutive_quota_errors >= stop_after_quota_errors:
                print(
                    f"Stopped after {consecutive_quota_errors} consecutive quota errors.",
                    flush=True,
                )
                break
            time.sleep(sleep_seconds)
        except Exception as error:
            print(f"[error] row {row_number}: {place_name}: {error}", flush=True)
            workbook.save(output_path)
            if is_quota_error(error):
                consecutive_quota_errors += 1
                if consecutive_quota_errors >= stop_after_quota_errors:
                    print(
                        f"Stopped after {consecutive_quota_errors} consecutive quota errors.",
                        flush=True,
                    )
                    break
            time.sleep(sleep_seconds)

        if should_fetch_youtube and youtube_urls:
            consecutive_quota_errors = 0

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"Saved: {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Fill IMAGE_URLS and YOUTUBE_URLS columns for attraction Excel data."
    )
    parser.add_argument("--input", required=True, help="Path to the source .xlsx file")
    parser.add_argument("--output", required=True, help="Path to the enriched .xlsx file")
    parser.add_argument(
        "--env",
        default="tools/media_enrichment.env",
        help="Path to env file with GOOGLE_API_KEY and GOOGLE_CSE_ID",
    )
    parser.add_argument("--limit", type=int, default=20, help="Rows to enrich. Use 0 for all rows.")
    parser.add_argument("--start-row", type=int, default=2, help="Excel row number to start from.")
    parser.add_argument("--overwrite", action="store_true", help="Replace existing media columns.")
    parser.add_argument("--sleep", type=float, default=0.25, help="Delay between rows.")
    parser.add_argument(
        "--media",
        choices=["all", "images", "youtube"],
        default="all",
        help="Which media type to enrich.",
    )
    parser.add_argument(
        "--stop-after-quota-errors",
        type=int,
        default=3,
        help="Stop after this many consecutive quota errors.",
    )
    parser.add_argument(
        "--youtube-query-mode",
        choices=["thai-only", "english-only", "all"],
        default="thai-only",
        help="Use fewer YouTube queries per row to save quota.",
    )
    parser.add_argument(
        "--youtube-provider",
        choices=["youtube-api", "custom-search"],
        default="youtube-api",
        help="Use YouTube Data API or Google Custom Search to find YouTube URLs.",
    )
    args = parser.parse_args()

    enrich_workbook(
        input_path=Path(args.input),
        output_path=Path(args.output),
        env_path=Path(args.env),
        limit=args.limit,
        start_row=args.start_row,
        overwrite=args.overwrite,
        sleep_seconds=args.sleep,
        media=args.media,
        stop_after_quota_errors=args.stop_after_quota_errors,
        youtube_query_mode=args.youtube_query_mode,
        youtube_provider=args.youtube_provider,
    )


if __name__ == "__main__":
    main()
